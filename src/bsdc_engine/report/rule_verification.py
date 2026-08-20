import sqlite3
from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.bsdc_engine.rules.store import RuleStore


def get_available_cu_ids(db_path: Path | None = None) -> list[str]:
    """Retrieve distinct non-global CU IDs stored in the rule_store database."""
    store = RuleStore(db_path=db_path)
    if not store.db_path.exists():
        return []
    with store.get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT DISTINCT cu_id FROM rule_store WHERE cu_id IS NOT NULL AND cu_id != '' AND is_global != 1"
        )
        rows = cursor.fetchall()
        return [r[0] for r in rows if r[0]]


def export_rule_verification_report(
    output_dir: Path,
    cu_id: str | None = None,
    db_path: Path | None = None,
) -> Path:
    """Export Rule Verification Report properly sorted according to Mapping file for a specific CU."""
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    store = RuleStore(db_path=db_path)

    if not cu_id:
        available_cus = get_available_cu_ids(db_path=db_path)
        cu_id = available_cus[0] if available_cus else "UNKNOWN"

    report_file = output_dir / f"Rule_Verification_Report_{cu_id}.xlsx"

    if not store.db_path.exists():
        print(f"❌ Database not found at {store.db_path}. Please run rule parsing first!")
        return report_file

    query = """
        SELECT 
            id AS "Rule_ID",
            cu_id AS "CU_ID",
            sheet_name AS "Sheet_Name",
            section_name AS "Section_Name",
            target_field AS "Target_Field",
            data_file AS "Source_File",
            column_letter AS "Source_Col",
            raw_notes AS "Raw_Notes",
            dsl_readable AS "Draft_Rule_DSL",
            rule_type AS "Rule_Type",
            parsed_by AS "Parsed_By",
            status AS "Current_Status"
        FROM rule_store
        WHERE cu_id = ? OR is_global = 1
        ORDER BY id ASC
    """

    with store.get_connection() as conn:
        df = pd.read_sql_query(query, conn, params=(cu_id,))

    if df.empty:
        print(f"⚠️ No Rule data in DB to export report for CU: [{cu_id}]!")
        return report_file

    # Add columns for QA assessment
    df["Decision (QA)"] = ""  # APPROVE / EDIT / REJECT
    df["QA Edited DSL"] = ""  # Input new DSL if EDIT is selected
    df["QA Reviewer"] = ""    # QA Reviewer Name
    df["QA Notes"] = ""       # QA Notes/Comments

    # Auto-suggest Decision for high-confidence rules
    df.loc[df["Current_Status"] == "AUTO_PARSED", "Decision (QA)"] = "APPROVE"
    df.loc[df["Current_Status"] == "NEEDS_REVIEW", "Decision (QA)"] = ""

    # Re-order columns for best QA visibility
    ordered_cols = [
        "Rule_ID",
        "Section_Name",
        "Target_Field",
        "Source_File",
        "Source_Col",
        "Raw_Notes",
        "Draft_Rule_DSL",
        "Decision (QA)",
        "QA Edited DSL",
        "QA Reviewer",
        "QA Notes",
        "Rule_Type",
        "Current_Status",
    ]
    df = df[ordered_cols]

    # Export to Excel and apply visual formatting
    with pd.ExcelWriter(report_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Rule_Verification", index=False)

        ws = writer.sheets["Rule_Verification"]

        # Header formatting
        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        # Section Rule row formatting
        section_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )
        section_font = Font(name="Calibri", size=10, bold=True, color="002060")

        # Format QA Decision cells with light yellow background
        qa_fill = PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )

        ws.freeze_panes = "A2"  # Freeze header row

        for col_num, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Coloring and aligning each row
        for row_idx, row_data in enumerate(df.itertuples(), start=2):
            target_field = str(row_data.Target_Field)
            is_section_rule = target_field == "_SECTION_RULE_"

            for col_idx in range(1, len(ordered_cols) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Highlight Decision & QA Edited cells
                col_name = ordered_cols[col_idx - 1]
                if col_name in ["Decision (QA)", "QA Edited DSL"]:
                    cell.fill = qa_fill

                # If Section Rule row -> Apply gray/blue section styling
                if is_section_rule:
                    cell.fill = section_fill
                    cell.font = section_font

        # Auto-adjust column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(
                max(max_len + 3, 12), 50
            )

    print("\n" + "=" * 60)
    print(f"📊 Successfully generated Verification Report for CU [{cu_id}]")
    print(f"👉 Location: {report_file.resolve()}")
    print("=" * 60 + "\n")
    return report_file