from datetime import datetime
from pathlib import Path
import openpyxl

from src.bsdc_engine.logging import get_logger
from src.bsdc_engine.validate.checks import (
    check_missing_data_file,
    check_converted_missing_mapping,
    check_empty_sheet,
)

logger = get_logger(__name__)


class MappingValidator:
    COL_FIELD = "Field"
    COL_DATA_FILE = "Data File"
    COL_COLUMN = "Column"
    COL_NOTES = "Notes/Additional Information"
    COL_CONVERTED = "Converted"

    def __init__(self, raw_dir: Path, output_report_dir: Path | None = None):
        self.raw_dir = Path(raw_dir)
        self.output_report_dir = Path(output_report_dir) if output_report_dir else None

        self.mapping_files = [
            f for f in self.raw_dir.glob("*.xlsx")
            if "mapping" in f.name.lower() and not f.name.startswith("~$")
        ]
        if not self.mapping_files:
            raise FileNotFoundError(
                f"❌ No Excel Mapping file (*mapping*.xlsx) found in: {self.raw_dir}"
            )

    def validate_single_file(self, file_path: Path) -> dict:
        errors_by_sheet = {}
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            return {"System": [f"Cannot open Excel Mapping file: {e}"]}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            col_idx_map = {}
            sheet_errors = []
            current_active_data_file = ""
            has_header = False
            mapped_field_count = 0

            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if not any(row):
                    continue

                row_str = [
                    str(cell).strip() if cell is not None else "" for cell in row
                ]

                # Identify Header row
                if self.COL_FIELD in row_str and self.COL_DATA_FILE in row_str:
                    col_idx_map = {
                        self.COL_FIELD: row_str.index(self.COL_FIELD),
                        self.COL_DATA_FILE: row_str.index(self.COL_DATA_FILE),
                        self.COL_COLUMN: (
                            row_str.index(self.COL_COLUMN)
                            if self.COL_COLUMN in row_str
                            else -1
                        ),
                        self.COL_NOTES: (
                            row_str.index(self.COL_NOTES)
                            if self.COL_NOTES in row_str
                            else -1
                        ),
                        self.COL_CONVERTED: (
                            row_str.index(self.COL_CONVERTED)
                            if self.COL_CONVERTED in row_str
                            else -1
                        ),
                    }
                    current_active_data_file = ""
                    has_header = True
                    continue

                if col_idx_map:
                    field_val = row_str[col_idx_map[self.COL_FIELD]]
                    if not field_val or field_val == self.COL_FIELD:
                        continue

                    data_file_val = row_str[col_idx_map[self.COL_DATA_FILE]]
                    column_val = (
                        row_str[col_idx_map[self.COL_COLUMN]]
                        if col_idx_map[self.COL_COLUMN] != -1
                        else ""
                    )
                    notes_val = (
                        row_str[col_idx_map[self.COL_NOTES]]
                        if col_idx_map[self.COL_NOTES] != -1
                        else ""
                    )
                    converted_val = (
                        row_str[col_idx_map[self.COL_CONVERTED]]
                        if col_idx_map[self.COL_CONVERTED] != -1
                        else ""
                    )

                    if data_file_val:
                        current_active_data_file = data_file_val
                    effective_data_file = data_file_val or current_active_data_file

                    is_converted = converted_val.lower() in [
                        "yes",
                        "y",
                        "true",
                        "1",
                    ]
                    has_source_mapping = bool(effective_data_file and column_val)
                    has_notes = bool(notes_val)

                    if is_converted or has_source_mapping or has_notes:
                        mapped_field_count += 1

                    # Preflight checks
                    err1 = check_missing_data_file(
                        row_idx, field_val, column_val, effective_data_file
                    )
                    if err1:
                        sheet_errors.append(err1)

                    err2 = check_converted_missing_mapping(
                        row_idx,
                        field_val,
                        is_converted,
                        has_source_mapping,
                        has_notes,
                    )
                    if err2:
                        sheet_errors.append(err2)

            # Sheet level check
            err3 = check_empty_sheet(has_header, mapped_field_count)
            if err3:
                sheet_errors.append(err3)

            if sheet_errors:
                errors_by_sheet[sheet_name] = sheet_errors

        wb.close()
        return errors_by_sheet

    def validate(self) -> tuple[bool, dict]:
        results_by_file = {}
        total_errors_all_files = 0

        for file in self.mapping_files:
            file_errors = self.validate_single_file(file)
            results_by_file[file.name] = file_errors
            total_errors_all_files += sum(
                len(errs) for errs in file_errors.values()
            )

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Generate Legacy Text Format Report
        if self.output_report_dir:
            self.output_report_dir.mkdir(parents=True, exist_ok=True)
            timestamp_file = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_file = (
                self.output_report_dir
                / f"Mapping_Preflight_Report_{timestamp_file}.txt"
            )

            lines = []
            lines.append("=" * 70)
            lines.append("MAPPING FILE VALIDATION REPORT (PREFLIGHT REPORT)")
            lines.append(f"Scan Time: {now_str}")
            lines.append(f"Total files scanned: {len(self.mapping_files)} file")
            lines.append(f"Total errors detected: {total_errors_all_files} errors")
            lines.append("=" * 70 + "\n")

            lines.append("📊 OVERALL STATUS OF FILES:")
            lines.append("-" * 70)
            for file_name, file_errs in results_by_file.items():
                file_err_cnt = sum(len(e) for e in file_errs.values())
                if file_err_cnt == 0:
                    lines.append(f"  • [{file_name}] -> ✔ 100% VALID")
                else:
                    lines.append(
                        f"  • [{file_name}] -> ❌ HAS ERRORS ({file_err_cnt} errors)"
                    )
            lines.append("=" * 70 + "\n")

            lines.append("🔍 DETAILS BY FILE MAPPING:\n")
            for file_name, file_errs in results_by_file.items():
                lines.append(f"📄 FILE MAPPING: [{file_name}]")
                lines.append("-" * 70)

                if not file_errs:
                    lines.append(
                        "✔ File is perfectly valid! No structural errors detected.\n"
                    )
                else:
                    for sheet, errs in file_errs.items():
                        lines.append(f"📑 Sheet: [{sheet}] ({len(errs)} errors)")
                        lines.append("-" * 70)
                        for err in errs:
                            lines.append(f"    ✕ {err}")
                        lines.append("")

            with open(report_file, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))

            logger.info(f"Generated QA Preflight Report at: {report_file}")

        is_all_valid = total_errors_all_files == 0
        return is_all_valid, results_by_file